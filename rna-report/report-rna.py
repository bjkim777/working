import os
import sys
import subprocess
import tempfile
import argparse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill

def getOpts():
	parser = argparse.ArgumentParser(description = '')
	parser.add_argument('-d', '--dir', type=str, required=True, metavar='<rna output dir>',  help= 'rna dir path')
	argv = parser.parse_args()
	return(argv)

def PANDAS_rnaBOX(data, index):
	GENE=[]
	VAL=[]
	SAMPLE=data.split('/')[-2]

	with open(data, 'r') as f:
		for i, line in enumerate(f):
			if i==0:
				pass
			else:
				col=line.split()
				GENE.append(col[0])
				VAL.append(col[3])

	if index==0:
		return pd.DataFrame(data=VAL, index=GENE, columns=[SAMPLE])
	else:
		return pd.Series(data=VAL, index=GENE)

### MAIN ###
def main (dir):
	FILE=[]
	command_line='find {0} -name  \'quant2.sf\' | sort'.format(dir)
	for line in subprocess.Popen(command_line, shell=True, stdout=subprocess.PIPE).stdout.readlines():
		FILE.append(line.strip().decode('utf-8'))

	for i, f in enumerate(FILE):
		SAMPLE=f.split('/')[-2]
		if i==0:
			BOX=PANDAS_rnaBOX(f, i)
		else:
			BOX[SAMPLE]=PANDAS_rnaBOX(f, i)

	BOX1=BOX.astype(float).round(0).astype(int)
	BOX2=np.log2(BOX.astype(float)).round(2)
	print(BOX1)
	print(BOX2)

	## EXCEL FORMAT ###
	EXCEL='asan-rna.xlsx'
	CELL_BOX = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )

	wb=openpyxl.Workbook()
	ws=wb.active
	ws.title='countsFromAbundance'

	for r in dataframe_to_rows(BOX1, header=True):
		ws.append(r)	

	# border style and column size
	COLUMN_WIDTHS=dict()
	for row in ws.iter_rows(max_col=len(list(ws.columns)), max_row=len(list(ws.rows)), min_row=1):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5

			# header style
			if cell.row==1 or cell.column==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

			# cell border
			cell.border=CELL_BOX
	
	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	ws1=wb.create_sheet()
	ws1.title='log2'

	for r in dataframe_to_rows(BOX2, header=True):
		ws1.append(r)

	# border style and column size
	COLUMN_WIDTHS=dict()
	for row in ws1.iter_rows(max_col=len(list(ws1.columns)), max_row=len(list(ws1.rows)), min_row=1):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5

			# header style
			if cell.row==1 or cell.column==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

			# cell border
			cell.border=CELL_BOX
	
	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws1.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	wb.save(EXCEL)
### ACTION ### 
if __name__ == "__main__":
	argv = getOpts()

	main( dir=argv.dir )
	sys.exit(0)