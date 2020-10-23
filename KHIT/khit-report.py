import os
import sys
import subprocess
import tempfile
import argparse
import sqlite3
import pandas as pd
import numpy as np
import openpyxl # v3.0.3
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill
from datetime import datetime
import time

def getOpts():
	parser = argparse.ArgumentParser(description = '')
	parser.add_argument('-e', '--excel', type=str, default='Kinase-hit_PDBs_details_v1.2_9_08_j.xlsx', metavar='<arg1>',  help= 'description')
	parser.add_argument('-d', '--db', type=str, default='khit-status.sql', metavar= '<sqlite db>', help= 'sqlite db')
	argv = parser.parse_args()
	return(argv)

def DF_sqlite(db):
	con=sqlite3.connect(db)
	cursor=con.cursor()

	cursor.execute("SELECT name FROM sqlite_master WHERE type=\'table\';")
	TABLE=[x[0] for x in cursor.fetchall()]

	df_t=pd.DataFrame()

	for table_name in TABLE:
		table=pd.read_sql('SELECT * FROM  "{0}"'.format(table_name), con=con)
		table.set_index("index",inplace=True)
		table.loc[table_name+"_total",:]=table.sum(axis=0)
		if len(list(df_t.columns))>0:
			for col in set(table.columns)-set(df_t.columns):
				table.drop(columns=[col], inplace=True)
		df_t=df_t.append(table)

	excel_to=df_t.loc[df_t.index.str.contains('total'),:]
	excel_to.index=[l.split('_')[:-1][0] for l in list(excel_to.index)]
	RESULT=excel_to.T.fillna(0)
	return RESULT.astype(int)

def DF_excel(excel):
	BOX=pd.read_excel(excel, header=1,index_col=0)
	start_index=int(list(BOX.columns).index('Blank')+1)
	end_index=int(list(BOX.columns).index('Suggestion'))
	
	hinge=BOX.T[start_index:end_index].T.values.tolist()
	hinge_filter=[[','.join([a for a in l if str(a)!='nan'])] for l in hinge]
	hinge_count=[[l[0], len(l[0].split(',')) if l[0]!=''  else 0] for l in hinge_filter]
	hinge_df=pd.DataFrame(hinge_count, columns=['Hinge', 'Hinge Count'])
	hinge_df.index+=1

	BOX['Lig']=BOX['Lig'].str.split('(', expand=True)

	return pd.merge(BOX[['Kinase', 'PDB', 'Lig','Chain','차수']], hinge_df, left_index=True, right_index=True, how='outer')

### MAIN ###
def main (db, excel):
	DF1=DF_excel(excel)
	DF2=DF_sqlite(db)
	print(DF1)
	print(DF2)

	KEY=(DF1['Kinase']+"_"+DF1['PDB'].str.lower()+DF1['Chain'].astype(str)+"_"+DF1['Lig'].astype(str))
	for key in KEY:
		if not 'CDK7' in key:
			KEY.replace(key, '_'.join(key.split('_')[:2]), inplace=True)

	DF1['KEY']=KEY
	MERGED_DB=pd.merge(DF1, DF2, how='right', left_on='KEY', right_index=True)
	
	MERGED_DB['postDM']=0
	MERGED_DB['preDM']=0
	MERGED_DB['preDM'][MERGED_DB['DM'] > 0]=1
	MERGED_DB['Complete']=''
	MERGED_DB['Server-info']='-'
	MERGED_DB['PDB(Lig)']=MERGED_DB['PDB']+"("+MERGED_DB['Lig'].astype(str)+")"

	MERGED_DB=MERGED_DB[['KEY', 'Kinase', 'PDB(Lig)', '차수', 'Hinge', 'Hinge Count', 'DARC', 'preDM', 'DM', 'postDM', 'Complete', 'Server-info']]
	print(MERGED_DB)

	# ##########################
	# # excel format
	# ##########################
	EXCEL='{}-khit-report.xlsx'.format(datetime.today().strftime('%y%m%d-%H%M%S'))
	CELL_BOX = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )

	wb=openpyxl.Workbook()
	ws=wb.active
	ws.title='khit report'

	for r in dataframe_to_rows(MERGED_DB, header=True, index=False):
		ws.append(r)

	COLUMN_WIDTHS=dict()
	for row in ws.iter_rows(max_col=len(list(ws.columns)), max_row=len(list(ws.rows)), min_row=1):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5

			# header style
			if cell.row==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

			# cell border
			cell.border=CELL_BOX

	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	wb.save(EXCEL)
### ACTION ### 
if __name__ == "__main__":
	argv = getOpts()

	main( db=argv.db , excel=argv.excel )
	sys.exit(0)