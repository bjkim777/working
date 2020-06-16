import os
import sys
import subprocess
import tempfile
import pandas as pd
import openpyxl # v3.0.3
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill
from datetime import datetime
import time

def getOpts():
	import argparse
	
	parser = argparse.ArgumentParser(description = '')
	parser.add_argument('-a', '--arg1', type=str, default='.', metavar='<arg1>',  help= 'description')
	parser.add_argument('-b', '--arg2', type=str, required=True,  metavar= '<arg2>', help= 'description')
	argv = parser.parse_args()
	return(argv)

class STinfo():
	
	CONFIG_FILE='~/.ssh/config'
	SQLITE_DB='/home/user1/miniconda3/cp-cmd.db'
	HEAD=['ID', 'HOST', 'INDEV', 'INMOUNT', 'OUTDEV', 'OUTMOUNT', 'OUTSN', 'STIME', 'ETIME', 'CMD']

	def __init__(self, IP):
		self.IP=IP
		self._ssh_client=None

	def INS_sshClient(self):
		import paramiko

		self._ssh_client=paramiko.SSHClient()
		ssh_config=paramiko.SSHConfig()
		
		self._ssh_client._policy=paramiko.WarningPolicy()
		self._ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

		user_config_file=os.path.expanduser(self.CONFIG_FILE)
		try:
			with open(user_config_file) as f:
				ssh_config.parse(f)
		except FileNotFoundError:
			print("{} file could not be found. Aborting.".format(user_config_file))
			sys.exit(1)
		
		user_config=ssh_config.lookup(self.IP)
		if 'proxycommand' in user_config:
			user_config['sock']=paramiko.ProxyCommand(user_config['proxycommand'])
			del user_config['proxycommand']
		if 'user' in user_config:
			user_config['username']=user_config['user']
			del user_config['user']
		print(user_config)

		self._ssh_client.connect(**user_config)
		return self._ssh_client

	def DF_sqlite3(self):
		client=self.INS_sshClient()

		stdin, stdout, stderr = client.exec_command('/home/user1/miniconda3/bin/sqlite3 {0} ".table"'.format(self.SQLITE_DB))
		table=stdout.readlines()[0].strip()
		stdin2, stdout2, stderr2 = client.exec_command('/home/user1/miniconda3/bin/sqlite3 -csv {0} "pragma table_info({1})"'.format(self.SQLITE_DB, table))
		HEAD=[line.strip().split(',')[1] for line in stdout2.readlines()]
		stdin3, stdout3, stderr3 = client.exec_command('/home/user1/miniconda3/bin/sqlite3 -csv {0} "select * from {1}"'.format(self.SQLITE_DB, table))
		BODY=list( \
			map(\
				lambda x : x[0:7]+[time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(x[7])))]+x[8:], \
				[line.strip().split(',') for line in stdout3.readlines()]))

		return pd.DataFrame(data=BODY, columns=HEAD)


### MAIN ###
def main ():
	BOX=pd.DataFrame(columns=STinfo.HEAD)
	IP=\
	['10.0.5.'+str(i) for i in [5,6,10]]+\
	['10.0.6.'+str(i) for i in [13,14]]+\
	['10.10.5.'+str(i) for i in [16]]+\
	['10.10.6.'+str(i) for i in [101, 102, 103, 104]]

	for ip in IP:
		# create object
		obj=STinfo(IP=ip)
		BOX=BOX.append(obj.DF_sqlite3(), ignore_index=True)
	
	##########################
	# excel format
	##########################
	EXCEL='{}-copy-status.xlsx'.format(datetime.today().strftime('%y%m%d-%H%M%S'))
	CELL_BOX = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )

	wb=openpyxl.Workbook()
	ws=wb.active
	ws.title='copy status'

	for r in dataframe_to_rows(BOX, header=True, index=False):
		ws.append(r)

	COLUMN_WIDTHS=dict()
	for row in ws.iter_rows(max_col=len(list(ws.columns)), max_row=len(list(ws.rows)), min_row=1):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5

			# header style
			if cell.row==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

			# cell border
			cell.border=CELL_BOX

	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	wb.save(EXCEL)


### ACTION ### 
if __name__ == "__main__":
	#argv = getOpts()

	main()
	sys.exit(0)
