import os
import sys
import subprocess
import tempfile
import pandas as pd
import numpy as np
import time
import argparse
import openpyxl # v3.0.3
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill
from datetime import datetime
from braceexpand import braceexpand
import paramiko

def getOpts():
	import argparse
	
	parser = argparse.ArgumentParser(description = '')
	parser.add_argument('-a', '--arg1', type=str, default='.', metavar='<arg1>',  help= 'description')
	parser.add_argument('-b', '--arg2', type=str, required=True,  metavar= '<arg2>', help= 'description')
	argv = parser.parse_args()
	return(argv)

def runTime(func):
	def wrapper(*args, **kwargs):
		start=time.time()
		result=func(*args, **kwargs)
		end=time.time()

		print ('# module : {0} / time(sec) : {1}'.format(func.__name__, end-start))

		return result
	return wrapper

class STinfo():
	
	CONFIG_FILE='~/.ssh/config'

	def __init__(self, IP):
		self.IP=IP
		self._ssh_client=None

	def INS_sshClient(self):
		import paramiko

		self._ssh_client=paramiko.SSHClient()
		ssh_config=paramiko.SSHConfig()
		
		self._ssh_client._policy=paramiko.WarningPolicy()
		self._ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

		user_config_file=os.path.expanduser(self.CONFIG_FILE)
		try:
			with open(user_config_file) as f:
				ssh_config.parse(f)
		except FileNotFoundError:
			print("{} file could not be found. Aborting.".format(user_config_file))
			sys.exit(1)
		
		user_config=ssh_config.lookup(self.IP)
		if 'proxycommand' in user_config:
			user_config['sock']=paramiko.ProxyCommand(user_config['proxycommand'])
			del user_config['proxycommand']
		if 'user' in user_config:
			user_config['username']=user_config['user']
			del user_config['user']
		print(user_config)

		self._ssh_client.connect(**user_config)
		return self._ssh_client

	@runTime
	def LIST_mountedClient(self):
		import time

		client=self.INS_sshClient()
		COUNT=10
		TOTAL=[]
		for i in range(1,COUNT+1):
			stdin, stdout, stderr = client.exec_command('netstat  | grep nfs | awk \'! /\.:/ {split($5,a,":"); print a[1]}\' | sort -V | uniq')
			TOTAL=list(set(TOTAL+[line.strip() for line in stdout.readlines()]))
			time.sleep(1)
		return TOTAL

	@runTime
	def LIST_mahaClient(self):
		client=self.INS_sshClient()
		stdin, stdout, stderr = client.exec_command('/usr/local/maha/gfs_lsclnt | sed 1d | awk \'{print $1}\'')
		return [line.strip() for line in stdout.readlines()]

def DF_ipBox(config):
	from braceexpand import braceexpand
	import pandas as pd

	HEADER=['HOST','IP','USER']
	VIEWER_BOX=pd.read_csv(config, sep='\t', header=None, names=HEADER)
	FULL_LIST_BOX=pd.DataFrame(columns=HEADER)

	with open(config) as f:
		for line in f:
			col=line.strip().split('\t')
			col=[w.replace('10.0', '10.1') for w in col]  # 10.0. -> 10.1.
			col=[w.replace('10.10', '10.11') for w in col]  # 10.10. -> 10.11.
			if '{' in col[1]:
				ip_list=list(braceexpand(col[1]))
				msg={HEADER[0]:[col[0]]*len(ip_list), HEADER[1]:ip_list, HEADER[2]:[col[2]]*len(ip_list)}
				df=pd.DataFrame(msg)
				FULL_LIST_BOX=pd.concat([FULL_LIST_BOX, df])
			else:	
				s=pd.Series(data=col, index=HEADER)
				FULL_LIST_BOX.loc[len(FULL_LIST_BOX.index)]=s
	
	FULL_LIST_BOX.index=FULL_LIST_BOX['IP'].tolist()
	VIEWER_BOX['IP']=VIEWER_BOX['IP'].str.replace('10.0', '10.1')
	VIEWER_BOX['IP']=VIEWER_BOX['IP'].str.replace('10.10', '10.11')
	VIEWER_BOX.index=VIEWER_BOX['IP'].tolist()

	return FULL_LIST_BOX, VIEWER_BOX


### MAIN ###
def main (conf=''):
	try:
		FULL_LIST_BOX, VIEWER_BOX=DF_ipBox(conf)
	except FileNotFoundError:
		print("\'{}\' file could not be found. Aborting.".format(conf))
		exit(1)

	STG_IP=['10.0.5.'+str(i) for i in [2,5,6,8,10,11,13]]+['10.0.6.3']

	for ip in STG_IP:
		obj=STinfo(IP=ip)

		VIEWER_BOX[ip]=0

		if '10.0.6' in ip:
			LIST=obj.LIST_mahaClient()
		else:
			LIST=obj.LIST_mountedClient()

		LIST=[ i for i in LIST if i.count('.') > 2 ]

		for l in LIST:
			if not l in FULL_LIST_BOX.index.tolist():
				FULL_LIST_BOX.loc[l]=['-',l,'-']+[np.nan]*(len(FULL_LIST_BOX.columns)-3)
				VIEWER_BOX.loc[l]=['-',l,'-']+[np.nan]*(len(VIEWER_BOX.columns)-3)
				VIEWER_BOX.loc[l:l,[ip]]=1
			else:
				if l in VIEWER_BOX.index.tolist():
					VIEWER_BOX.loc[l:l, [ip]]=1
				else:
					for v in VIEWER_BOX.index.tolist():
						if '{' in v:
							if l in list(braceexpand(v)):
								if str(VIEWER_BOX.loc[v:v,[ip]].values.tolist()[0][0])=='nan':
									VIEWER_BOX.loc[v:v,[ip]]=1
								else:
									VIEWER_BOX.loc[v:v,[ip]]=VIEWER_BOX.loc[v:v,[ip]].add(1)

		FULL_LIST_BOX[ip] = pd.Series(data=[1]*len(LIST), index=LIST)

	FULL_LIST_BOX.fillna(0, inplace=True)
	VIEWER_BOX.fillna(0, inplace=True)

	##########################
	# excel format
	##########################
	EXCEL='{}-mounted_client.xlsx'.format(datetime.today().strftime('%y%m%d-%H%M%S'))
	CELL_BOX = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )

	wb=openpyxl.Workbook()
	ws0=wb.active
	ws0.title='sum'
	SUM=pd.DataFrame(VIEWER_BOX.sum(), columns=['client'])[3:].T

	for r in dataframe_to_rows(SUM, header=True, index=True):
		ws0.append(r)
	ws0.delete_rows(2)

	for row in ws0.iter_rows(max_col=len(list(ws0.columns)), max_row=len(list(ws0.rows)), min_row=1):
		for i, cell in enumerate(row):
			cell.border=CELL_BOX
			if cell.row==1 or cell.column==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

	ws=wb.create_sheet()
	ws.title='show client'

	# dataframe to excel
	for r in dataframe_to_rows(VIEWER_BOX, index=False, header=True):
		ws.append(r)

	# cell style
	COLUMN_WIDTHS=dict()
	for row in ws.iter_rows(max_col=len(list(ws.columns)), max_row=len(list(ws.rows)), min_row=1):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5

			# header style
			if cell.row==1:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro

			# check connected 
			if cell.row!=1 and cell.column>3 and cell.value>0:
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFFFFF00')) # color : Yellow

			# cell border
			cell.border=CELL_BOX

	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	ws2=wb.create_sheet()
	ws2.title='full'
	for r in dataframe_to_rows(FULL_LIST_BOX, index=False, header=True):
		ws2.append(r)

	wb.save(EXCEL)


### ACTION ### 
if __name__ == "__main__":
	#argv = getOpts()
	CONFIG_FILE='host-server-user-200527.txt'

	main(conf=CONFIG_FILE)
	sys.exit(0)
