import os
import sys
import subprocess
import tempfile
import pandas as pd
import openpyxl # v3.0.3
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill
from datetime import datetime
from natsort import natsorted

def getOpts():
	import argparse
	
	parser = argparse.ArgumentParser(description = '')
	parser.add_argument('-a', '--arg1', type=str, default='.', metavar='<arg1>',  help= 'description')
	parser.add_argument('-b', '--arg2', type=str, required=True,  metavar= '<arg2>', help= 'description')
	argv = parser.parse_args()
	return(argv)

class STinfo():
	
	TOTAL_HEAD=['Storage', 'Size(TB)', 'Used(TB)', 'Available(TB)', 'Use%','Mounted', \
	'Data List', 'Manager', 'Expiration Date', 'Backup/Description']
	
	HEAD=['Storage', 'Size(TB)', 'Used(TB)', 'Available(TB)', 'Use%','Mounted','Data List']
	CONFIG_FILE='~/.ssh/config'

	def __init__(self, IP):
		self.IP=IP
		self._ssh_client=None

	def LIST_showFolder(self, FOLDER):
		from random import shuffle

		DIR=[ls for ls in os.listdir(FOLDER) if not os.path.isfile(os.path.join(FOLDER, ls))]
		shuffle(DIR)
		if len(DIR)>4:
			DIR=DIR[:4]
		return DIR

	def INS_sshClient(self):
		import paramiko

		self._ssh_client=paramiko.SSHClient()
		ssh_config=paramiko.SSHConfig()
		
		self._ssh_client._policy=paramiko.WarningPolicy()
		self._ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

		user_config_file=os.path.expanduser(self.CONFIG_FILE)
		try:
			with open(user_config_file) as f:
				ssh_config.parse(f)
		except FileNotFoundError:
			print("{} file could not be found. Aborting.".format(user_config_file))
			sys.exit(1)
		
		user_config=ssh_config.lookup(self.IP)
		if 'proxycommand' in user_config:
			user_config['sock']=paramiko.ProxyCommand(user_config['proxycommand'])
			del user_config['proxycommand']
		if 'user' in user_config:
			user_config['username']=user_config['user']
			del user_config['user']
		print(user_config)

		self._ssh_client.connect(**user_config)
		return self._ssh_client

	def LIST_selectST(self):
		client=self.INS_sshClient()
		stdin, stdout, stderr = client.exec_command('df -l \
					| awk \' !/maha/ {if(length($2)> 9 && NR!=1) \
					print substr($0,index($0,$2))}\'')

		TOTAL_LIST=[]
		LEN=5
		for line in [line.strip().split() for line in stdout.readlines()]:
			stdin2, stdout2, stderr2 = client.exec_command('find %s -maxdepth 1 -type d ! -name \'lost+found\' \
				| sed 1d \
				|awk \'{split($0,a,"/"); print a[length(a)]}\''% line[-1])

			all_list=[a.strip() for a in stdout2.readlines()]
			if len(all_list)>=LEN:
				all_list=all_list[:LEN-1]+['...']

			line.append(', '.join(all_list))
			TOTAL_LIST.append(line)

		return TOTAL_LIST

	def LIST_mahaST(self):
		TOTAL_LIST=[]
		client=self.INS_sshClient()
		stdin, stdout, stderr = client.exec_command('/usr/local/maha/gfs_lsvol -l | awk \'/maha/ {print $(NF-1),$(NF-2),$1}\'')
		for line in [line.strip().split() for line in stdout.readlines()]:
			trans=[]
			for s in line[0:2]:
				if 'T' in s:
					mu=float(s[:-1])*10**9
					trans.append(mu)
				elif 'G' in s:
					mu=float(s[:-1])*10**6
					trans.append(mu)
				elif 'M' in s:
					mu=float(s[:-1])*10**3
					trans.append(mu)
				elif 'K' in s:
					mu=float(s[:-1])
					trans.append(mu)
				elif 'B' in s:
					mu=float(s[:-1])*10**-3
					trans.append(mu)
				else:
					trans.append(s)

			TOTAL_LIST.append([trans[0],trans[1],trans[0]-trans[1],str(trans[1]/trans[0]*100)+'%',line[-1],''])
		return TOTAL_LIST

### MAIN ###
def main ():
	BOX=pd.DataFrame(index=STinfo.TOTAL_HEAD)

	DIC={#'10.0.5.2':'#100-02', #'10.0.5.3':'#100-03', '10.0.5.4':'#100-04',
		'10.0.5.5':'#100-05', 
		'10.0.5.6':'#160-06', '10.0.5.10':'#500-10',
		'10.0.2.205':'#D205', '10.0.2.206':'#D206', '10.0.2.207':'#D207', 
		'10.0.2.208':'#D208', '10.0.2.209':'#D209', '10.0.2.210':'#D210', 
		'10.0.6.3':'#MAHA'}
	IP=list(DIC.keys())
	natsorted(IP)
	INDEX=1

	for ip in IP:
		# create object
		obj=STinfo(IP=ip)
		
		# extract storage information
		if 'MAHA' in DIC[ip]:
			for stinfo in obj.LIST_mahaST():
				stinfo.insert(0,DIC[ip])
				line=pd.Series(stinfo, index=STinfo.HEAD)
				BOX[INDEX]=line
				INDEX+=1
		else:
			for stinfo in obj.LIST_selectST():
				stinfo.insert(0,DIC[ip])
				line=pd.Series(stinfo, index=STinfo.HEAD)
				BOX[INDEX]=line
				INDEX+=1
		
		# change type 'float'
		BOX=BOX.T.astype({'Size(TB)':'float','Used(TB)':'float','Available(TB)':'float'})
		BOX=BOX.T

		# change 'sum cell' information
		SUM=BOX.T[BOX.T.Storage==DIC[ip]].sum()
		for col in ['Storage', 'Use%', 'Mounted', "Data List",'Manager', 'Expiration Date', 'Backup/Description']:
			if col=='Storage': 
				SUM[col]=DIC[ip]
			else:
				SUM[col]='-'
		BOX[INDEX]=SUM
		INDEX+=1

	# extract size unit (KB -> TB)
	for sz in ['Size(TB)', 'Used(TB)', 'Available(TB)']:
		BOX.loc[sz]=BOX.loc[sz]/pow(10,9)
	
	##########################
	# excel format
	##########################
	EXCEL='{}-storage-stat.xlsx'.format(datetime.today().strftime('%y%m%d'))
	CELL_BOX = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )

	wb=openpyxl.Workbook()
	ws=wb.active
	ws.title='Storage'

	# dataframe to excel
	for r in dataframe_to_rows(BOX.T, index=False, header=True):
		ws.append(r)
	ws.insert_rows(1, amount=2)

	# border style and column size
	COLUMN_WIDTHS=dict()
	for row in ws.iter_rows(max_col=len(list(ws.columns)), max_row=len(list(ws.rows)), min_row=3):
		for i, cell in enumerate(row) :
			# column size
			if i+1 in COLUMN_WIDTHS:
				if COLUMN_WIDTHS[i+1]<len(str(cell.value))+5:
					COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			else:
				COLUMN_WIDTHS[i+1]=len(str(cell.value))+5
			# 'sum' cell style
			if cell.row==3:
				cell.font=Font(bold=True)
				cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDCDC')) # color : Gainsboro
			cell.border=CELL_BOX

	# adjust auto column size
	for key in list(COLUMN_WIDTHS.keys()):
		ws.column_dimensions[get_column_letter(key)].width=COLUMN_WIDTHS[key]

	# merge cells
	for ip in IP:
		merge_list=list()
		for row in ws.rows:
			if row[0].value==DIC[ip]:
				merge_list.append(row[0].row)
		ws.merge_cells(start_row=merge_list[0], end_row=merge_list[-1], start_column=1, end_column=1)
		ws.cell(row=merge_list[0], column=1).alignment=Alignment(horizontal='center', vertical='center')
		ws.cell(row=merge_list[0], column=1).font=Font(bold=True, color=Color(rgb='FF1E90FF')) # color : DodgerBlue
		for col_index in range(2,len(list(ws.columns))+1):
			ws.cell(row=merge_list[-1], column=col_index).fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFF8F8FF')) # color : GhostWhite
			ws.cell(row=merge_list[-1], column=col_index).font=Font(bold=True, color=colors.RED)

	# description part
	ws['A2']='설명 : #100-02 [100TB / IP 10.1.5.xx], #D205 [개발서버 IP 205]'
	ws['A2'].font=Font(bold=True, color=Color(rgb='FF87CEEB')) # color : SkyBlue
	ws['J2']='Update: {}'.format(datetime.today().strftime('%Y.%m.%d'))
	ws['J2'].font=Font(bold=True, color=colors.RED)
	wb.save(EXCEL)

### ACTION ### 
if __name__ == "__main__":
	#argv = getOpts()

	main()
	sys.exit(0)
